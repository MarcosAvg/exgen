import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.catalog_models import EvidencePhotoData, CatalogSystem
from src.utils.config_manager import get_catalog_system


class ExcelRegistryService:
    """Servicio para gestionar el inventario completo de evidencias en Excel."""

    DEFAULT_FILENAME = "Registros.xlsx"
    
    MESES = [
        "", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]

    def __init__(self, base_dir: str):
        self.output_path = os.path.join(base_dir, self.DEFAULT_FILENAME)
        self.catalog_system = get_catalog_system()

        # Paleta de colores profesionales
        self.COLORS = {
            "banner_bg": "1B365D",
            "group_inspection": "4338CA",
            "group_results": "0D9488",
            "header_inspection": "4F46E5",
            "header_results": "14B8A6",
            "zebra": "F8FAFC",
            "border": "CBD5E1",
            "white": "FFFFFF",
            "done_fg": "15803D",
            "done_bg": "DCFCE7",
            "pending_fg": "92400E", # Marrón/Naranja oscuro
            "pending_bg": "FEF3C7", # Amarillo muy claro
            "pdf_bg": "BE123C",
            "cell_inspec_1": "F9FAFB",
            "cell_inspec_2": "EEF2FF",
            "cell_result_1": "F9FAFB",
            "cell_result_2": "F0FDFA",
        }

        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS["border"]),
            right=Side(style='thin', color=self.COLORS["border"]),
            top=Side(style='thin', color=self.COLORS["border"]),
            bottom=Side(style='thin', color=self.COLORS["border"])
        )

        self.styles = {
            "banner": {
                "font": Font(name="Arial", size=16, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["banner_bg"], end_color=self.COLORS["banner_bg"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "meta": {
                "font": Font(name="Arial", size=10, bold=True, color="444444"),
                "alignment": Alignment(horizontal="left", vertical="center"),
            },
            "group_header_inspection": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["group_inspection"], end_color=self.COLORS["group_inspection"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "group_header_results": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["group_results"], end_color=self.COLORS["group_results"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "header_inspection": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["header_inspection"], end_color=self.COLORS["header_inspection"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "header_results": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["header_results"], end_color=self.COLORS["header_results"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "cell_normal": {
                "font": Font(name="Arial", size=10),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "status_done": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["done_fg"]),
                "fill": PatternFill(start_color=self.COLORS["done_bg"], end_color=self.COLORS["done_bg"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "status_pending": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["pending_fg"]),
                "fill": PatternFill(start_color=self.COLORS["pending_bg"], end_color=self.COLORS["pending_bg"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            },
            "pdf_link": {
                "font": Font(name="Arial", size=10, bold=True, color=self.COLORS["white"]),
                "fill": PatternFill(start_color=self.COLORS["pdf_bg"], end_color=self.COLORS["pdf_bg"], fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self.thin_border,
            }
        }

    def _get_dynamic_header(self, data: EvidencePhotoData) -> str:
        if not data.dependent_values:
            return "DETALLE"
        first_key = list(data.dependent_values.keys())[0]
        cat = self.catalog_system.get_catalog_by_name(first_key)
        return cat.label.upper() if cat else "DETALLE"

    def _get_safe_sheet_name(self, equipo: str, edificio: str = None) -> str:
        if not edificio: return equipo[:31]
        name = f"{equipo[:14]} - {edificio[:14]}"
        return name[:31]

    def _get_month_period(self, wb: Workbook) -> str:
        dates = set()
        for ws in wb.worksheets:
            for row in range(6, ws.max_row + 1):
                val = ws.cell(row=row, column=2).value # Col B
                if val and isinstance(val, str) and "-" in val and val != "Pendiente":
                    try:
                        parts = val.split("-")
                        if len(parts) == 3:
                            month, year = int(parts[1]), int(parts[2])
                            if year < 100: year += 2000
                            dates.add((year, month))
                    except: continue
        if not dates: return "N/A"
        sorted_dates = sorted(list(dates))
        min_d, max_d = sorted_dates[0], sorted_dates[-1]
        if min_d == max_d: return f"{self.MESES[min_d[1]]} {min_d[0]}"
        elif min_d[0] == max_d[0]: return f"{self.MESES[min_d[1]]} - {self.MESES[max_d[1]]} {min_d[0]}"
        else: return f"{self.MESES[min_d[1]]} {min_d[0]} - {self.MESES[max_d[1]]} {max_d[0]}"

    def _setup_new_sheet(self, ws, equipo_name, detail_label):
        """Dibuja encabezados básicos."""
        ws.merge_cells('A1:F2')
        ws['A1'].value = "REGISTRO DE EVIDENCIAS - EXTER"
        ws['A1'].font, ws['A1'].fill, ws['A1'].alignment = self.styles["banner"]["font"], self.styles["banner"]["fill"], self.styles["banner"]["alignment"]

        if 'A3:F3' not in ws.merged_cells: ws.merge_cells('A3:F3')
        
        ws.merge_cells('A4:C4')
        ws['A4'].value = "DATOS DE INSPECCIÓN"
        for col in range(1, 4):
            c = ws.cell(row=4, column=col)
            c.font, c.fill, c.alignment, c.border = self.styles["group_header_inspection"]["font"], self.styles["group_header_inspection"]["fill"], self.styles["group_header_inspection"]["alignment"], self.styles["group_header_inspection"]["border"]

        ws.merge_cells('D4:F4')
        ws['D4'].value = "RESULTADOS Y EVIDENCIA"
        for col in range(4, 7):
            c = ws.cell(row=4, column=col)
            c.font, c.fill, c.alignment, c.border = self.styles["group_header_results"]["font"], self.styles["group_header_results"]["fill"], self.styles["group_header_results"]["alignment"], self.styles["group_header_results"]["border"]

        ws.row_dimensions[5].height = 25
        headers = ["UBICACIÓN", "FECHA", detail_label.upper(), "ESTADO", "EVIDENCIA", "NOTAS"]
        for col, text in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=text)
            key = "header_inspection" if col <= 3 else "header_results"
            c.font, c.fill, c.alignment, c.border = self.styles[key]["font"], self.styles[key]["fill"], self.styles[key]["alignment"], self.styles[key]["border"]

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25

    def _sync_all_meta_rows(self, wb: Workbook, current_building: str):
        periodo = self._get_month_period(wb)
        for ws in wb.worksheets:
            # Intentar extraer edificio del título (EQUIPO - EDIFICIO) o de la primera fila de datos
            title_parts = ws.title.split(" - ")
            equipo = title_parts[0].upper()
            
            # Detectar ubicación específica de la hoja
            ubicacion = None
            if len(title_parts) > 1:
                ubicacion = title_parts[1].upper()
            else:
                # Si no está en el título, buscar el primer valor en la columna A (Ubicación)
                first_row_val = ws.cell(row=6, column=1).value
                if first_row_val:
                    ubicacion = str(first_row_val).upper()
            
            # Fallback al edificio actual si no se detectó nada
            if not ubicacion:
                ubicacion = current_building.upper()

            txt = f"UBICACIÓN: {ubicacion}  |  EQUIPO: {equipo}  |  PERIODO: {periodo}"
            ws['A3'].value = txt
            ws['A3'].font, ws['A3'].alignment, ws['A3'].fill = self.styles["meta"]["font"], self.styles["meta"]["alignment"], PatternFill(start_color="FDFDFD", end_color="FDFDFD", fill_type="solid")
            if 'A3:F3' not in ws.merged_cells:
                try: ws.merge_cells('A3:F3')
                except: pass

    def sync_catalogs(self, catalog_system: CatalogSystem):
        """Pre-puebla el inventario completo basado en los catálogos."""
        wb = None
        if os.path.exists(self.output_path):
            try: wb = load_workbook(self.output_path)
            except: wb = Workbook()
        else:
            wb = Workbook()

        edificios = catalog_system.get_base_catalogs()[0].items
        tipos_equipo = catalog_system.get_base_catalogs()[1].items
        dept_cats = catalog_system.get_dependent_catalogs()

        for equipo in tipos_equipo:
            for edif in edificios:
                # Buscar catálogos activos para este Edif/Equipo (Lógica AND)
                detail_cat = None
                for cat in dept_cats:
                    # Si un catálogo no tiene dependencias, aplica a todas las combinaciones
                    if not cat.dependencies:
                        detail_cat = cat
                        break
                    
                    all_met = True
                    for dep in cat.dependencies:
                        if dep.parent_name == "edificio":
                            if edif not in dep.values:
                                all_met = False; break
                        elif dep.parent_name == "tipo_equipo":
                            if equipo not in dep.values:
                                all_met = False; break
                        # Nota: Para pre-población, solo consideramos dependencias de base (Edif/Equipo)
                    
                    if all_met:
                        detail_cat = cat
                        break
                
                # SI NO HAY MATCH Y EXISTEN CATÁLOGOS: Saltamos esta combinación (Evita hojas/filas fantasma)
                if not detail_cat and dept_cats:
                    continue

                details = detail_cat.items if detail_cat else [""]
                detail_label = detail_cat.label if detail_cat else "DETALLE"
                
                # Nombre de hoja inteligente (un solo edificio -> Equipo, múltiple -> Equipo - Edif)
                # Pero para pre-población masiva, si el equipo está en múltiples edificios del catálogo, usamos sufijo.
                other_edifs = [e for e in edificios if e != edif]
                has_multi = False
                for e in other_edifs:
                    for cat in dept_cats:
                        if cat.parent_name == "edificio" and e in cat.parent_values:
                            has_multi = True; break
                
                sheet_name = self._get_safe_sheet_name(equipo, edif if has_multi else None)
                if sheet_name in wb.sheetnames: ws = wb[sheet_name]
                else: 
                    ws = wb.create_sheet(sheet_name)
                    if "Sheet" in wb.sheetnames: del wb["Sheet"]
                    self._setup_new_sheet(ws, equipo, detail_label)
                
                # Poblar sub-detalles (pisos, etc)
                for det in details:
                    # Buscar si ya existe la fila
                    found = False
                    for r in range(6, ws.max_row + 1):
                        if ws.cell(row=r, column=1).value == edif and ws.cell(row=r, column=3).value == det:
                            found = True; break
                    
                    if not found:
                        idx = ws.max_row + 1
                        ws.row_dimensions[idx].height = 25
                        is_even = (idx % 2 == 0)
                        
                        # Ubicación, Fecha, Detalle
                        ws.cell(row=idx, column=1, value=edif)
                        ws.cell(row=idx, column=2, value="Pendiente")
                        ws.cell(row=idx, column=3, value=det)
                        
                        for col in range(1, 4):
                            c = ws.cell(row=idx, column=col)
                            c.font = self.styles["cell_normal"]["font"]
                            bg = self.COLORS["cell_inspec_2"] if is_even else self.COLORS["cell_inspec_1"]
                            c.fill, c.alignment, c.border = PatternFill(start_color=bg, end_color=bg, fill_type="solid"), self.styles["cell_normal"]["alignment"], self.styles["cell_normal"]["border"]
                        
                        # Estado: Sin evidencia
                        cs = ws.cell(row=idx, column=4, value="Sin evidencia")
                        cs.font, cs.fill, cs.alignment, cs.border = self.styles["status_pending"]["font"], self.styles["status_pending"]["fill"], self.styles["status_pending"]["alignment"], self.styles["status_pending"]["border"]
                        
                        # Placeholder PDF y Notas
                        ws.cell(row=idx, column=5, value="-")
                        ws.cell(row=idx, column=6, value="")
                        for col in (5, 6):
                            c = ws.cell(row=idx, column=col)
                            bg = self.COLORS["cell_result_2"] if is_even else self.COLORS["cell_result_1"]
                            c.font, c.fill, c.alignment, c.border = self.styles["cell_normal"]["font"], PatternFill(start_color=bg, end_color=bg, fill_type="solid"), self.styles["cell_normal"]["alignment"], self.styles["cell_normal"]["border"]

        self._sync_all_meta_rows(wb, edificios[0] if edificios else "PROYECTO")
        wb.save(self.output_path)

    def update_registry(self, data: EvidencePhotoData, pdf_path: str):
        wb = None
        if os.path.exists(self.output_path):
            try: wb = load_workbook(self.output_path)
            except: wb = Workbook()
        else:
            wb = Workbook()

        # Determinar nombre de hoja real (buscando si existen otras ubicaciones en el archivo)
        prefix = f"{data.tipo_equipo[:14]} - "
        has_multi = any(s.startswith(prefix) for s in wb.sheetnames)
        if not has_multi:
            # Si no hay con guión, verificamos si existe la simple pero con otro edificio
            if data.tipo_equipo in wb.sheetnames:
                ws_simple = wb[data.tipo_equipo]
                first_b = ws_simple.cell(row=6, column=1).value
                if first_b and first_b != data.edificio:
                    has_multi = True
        
        target_name = self._get_safe_sheet_name(data.tipo_equipo, data.edificio if has_multi else None)
        
        # Si hubo cambio de "simple" a "multi" justo ahora
        if has_multi and data.tipo_equipo in wb.sheetnames:
            ws_simple = wb[data.tipo_equipo]
            first_b = ws_simple.cell(row=6, column=1).value
            ws_simple.title = self._get_safe_sheet_name(data.tipo_equipo, first_b)

        if target_name in wb.sheetnames: ws = wb[target_name]
        else:
            ws = wb.create_sheet(target_name)
            if "Sheet" in wb.sheetnames: del wb["Sheet"]
            detail_label = self._get_dynamic_header(data)
            self._setup_new_sheet(ws, data.tipo_equipo, detail_label)

        # Buscar fila existente (Ubicación + Detalle)
        target_det = list(data.dependent_values.values())[0] if data.dependent_values else ""
        target_idx = -1
        for r in range(6, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == data.edificio and ws.cell(row=r, column=3).value == target_det:
                target_idx = r; break
        
        if target_idx == -1:
            target_idx = ws.max_row + 1
            ws.row_dimensions[target_idx].height = 25
        
        is_even = (target_idx % 2 == 0)
        
        # Actualizar Datos
        ws.cell(row=target_idx, column=1, value=data.edificio)
        ws.cell(row=target_idx, column=2, value=data.fecha)
        ws.cell(row=target_idx, column=3, value=target_det)
        
        for col in range(1, 4):
            c = ws.cell(row=target_idx, column=col)
            c.font = self.styles["cell_normal"]["font"]
            bg = self.COLORS["cell_inspec_2"] if is_even else self.COLORS["cell_inspec_1"]
            c.fill, c.alignment, c.border = PatternFill(start_color=bg, end_color=bg, fill_type="solid"), self.styles["cell_normal"]["alignment"], self.styles["cell_normal"]["border"]

        # Estatus Completado
        cs = ws.cell(row=target_idx, column=4, value="Completado")
        cs.font, cs.fill, cs.alignment, cs.border = self.styles["status_done"]["font"], self.styles["status_done"]["fill"], self.styles["status_done"]["alignment"], self.styles["status_done"]["border"]

        # Link PDF
        rel = os.path.relpath(pdf_path, os.path.dirname(self.output_path))
        cl = ws.cell(row=target_idx, column=5, value="PDF")
        cl.hyperlink = rel
        cl.font, cl.fill, cl.alignment, cl.border = self.styles["pdf_link"]["font"], self.styles["pdf_link"]["fill"], self.styles["pdf_link"]["alignment"], self.styles["pdf_link"]["border"]

        # Notas (limpiar placeholder si había)
        cn = ws.cell(row=target_idx, column=6, value="")
        cn.font = self.styles["cell_normal"]["font"]
        bg_n = self.COLORS["cell_result_2"] if is_even else self.COLORS["cell_result_1"]
        cn.fill, cn.alignment, cn.border = PatternFill(start_color=bg_n, end_color=bg_n, fill_type="solid"), self.styles["cell_normal"]["alignment"], self.styles["cell_normal"]["border"]

        self._sync_all_meta_rows(wb, data.edificio)
        wb.save(self.output_path)
        return self.output_path

    def check_registry_status(self, data: EvidencePhotoData) -> bool:
        """
        Verifica si la combinación actual ya está marcada como Completada en el Excel.
        """
        if not os.path.exists(self.output_path):
            return False
            
        try:
            wb = load_workbook(self.output_path, read_only=True)
            
            # Misma lógica de resolución de nombre de hoja que en update_registry
            prefix = f"{data.tipo_equipo[:14]} - "
            has_multi = any(s.startswith(prefix) for s in wb.sheetnames)
            if not has_multi:
                if data.tipo_equipo in wb.sheetnames:
                    ws_simple = wb[data.tipo_equipo]
                    first_b = ws_simple.cell(row=6, column=1).value
                    if first_b and first_b != data.edificio:
                        has_multi = True
            
            target_name = self._get_safe_sheet_name(data.tipo_equipo, data.edificio if has_multi else None)
            
            if target_name not in wb.sheetnames:
                wb.close()
                return False
                
            ws = wb[target_name]
            target_det = list(data.dependent_values.values())[0] if data.dependent_values else ""
            
            is_completed = False
            for r in range(6, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == data.edificio and ws.cell(row=r, column=3).value == target_det:
                    estado = ws.cell(row=r, column=4).value
                    if estado == "Completado":
                        is_completed = True
                    break
                    
            wb.close()
            return is_completed
        except Exception:
            return False
